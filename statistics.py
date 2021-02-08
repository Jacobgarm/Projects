import itertools
import math
from dataclasses import dataclass
import matplotlib.pyplot as plt
from decimal import Decimal
from numpy.lib.arraysetops import isin
import openpyxl
import numpy as np
from itertools import combinations

@dataclass
class linear:
    a: Decimal
    b: Decimal

    def __call__(self, x):
        return self.a * x + self.b

@dataclass
class exponential:
    a: Decimal
    b: Decimal

    def __call__(self, x):
        return self.b * self.a**x

@dataclass
class power:
    a: Decimal
    b: Decimal

    def __call__(self, x):
        return self.b * x**self.a

@dataclass
class polynomial:
    ps: list[Decimal]

    def __call__(self, x):
        return sum([self.ps[i] * x ** ((len(self.ps) - 1) - i) for i in range(len(self.ps))])

    def degree(self):
        return len(self.p)-1

@dataclass
class multiLinear:
    ps: list[Decimal]

    def __call__(self, *xs):
        return self.ps[0] + sum(self.ps[i+1] * xs[i] for i in range(len(xs)))

    def degree(self):
        return len(self.p)-1

@dataclass
class logarithmic:
    a: Decimal
    b: Decimal

    def __call__(self, x):
        return self.a * x.ln() + self.b



def apply(X, func):
    return [func(x) for x in X]

# Single set descriptors

def mean(X):
    return sum(X) / len(X)

def percentile(X, q):
    S = sorted(X)
    i = q * (len(X) - 1)
    if i % 1 == 0:
        return S[int(i)]
    return S[int(i)] + Decimal(q) * (S[int(i) + 1] - S[int(i)])

def median(X):
    return percentile(X, 0.5)

def mode(X):
    return max(X, key = X.count)

def geometricMean(X):
    return math.prod(X) ** (1 / Decimal(len(X)))

def harmonicMean(X):
    return 1 / (sum(1 / x for x in X) / len(X))

def variation(X):
    return max(X) - min(X)

def meanAbsoluteDeviation(X):
    m = mean(X)
    return mean([abs(x - m) for x in X])

def sumSquares(X):
    return sum(x**2 for x in X)

def variance(X):
    m = mean(X)
    squareDeviations = [(x - m) ** 2 for x in X]
    return mean(squareDeviations)

def standardDeviation(X):
    return variance(X).sqrt()

# Dual set descriptors

def sumProducts(X, Y):
    assert len(X) == len(Y)
    return sum([X[i] * Y[i] for i in range(len(X))])

# Model tests

def residuals(X, Y, model):
    if isinstance(X, Decimal):
        assert len(X) == len(Y)
        return [Y[i] - model(X[i]) for i in range(len(X))]
    else:
        return [Y[i] - model(*[p[i] for p in X]) for i in range(len(Y))]

def r2(X, Y, model):
    my = mean(Y)
    tot = sumSquares([y - my for y in Y])
    res = sumSquares(residuals(X, Y, model))
    return 1 - res/tot

# Regresseion

def conReg(X,Y):
    assert len(X) == len(Y)
    c = mean(Y)
    return linear(Decimal(0), c)

def proReg(X, Y):
    assert len(X) == len(Y)
    p = sumProducts(X, Y) / sumSquares(X)
    return linear(p, Decimal(0))

def linReg(X, Y):
    assert len(X) == len(Y)
    n = len(X)
    Sx = sum(X)
    Sy = sum(Y)
    Sxy = sumProducts(X, Y)
    Sxx = sumSquares(X)
    a = (n*Sxy - Sx*Sy) / (n*Sxx - Sx**2)
    b = (Sy * Sxx - Sx * Sxy) / (n * Sxx - Sx**2)
    return linear(a, b)

def theilSen(X, Y):
    assert len(X) == len(Y)
    slopes = []
    for i, j in combinations(list(range(len(X))), 2):
        if X[i] == X[j]:
            continue
        slopes.append((Y[i] - Y[j]) / (X[i] - X[j]))
    m = median(slopes)
    intercepts = []
    for i in range(len(X)):
        intercepts.append(Y[i] - m * X[i])
    b = median(intercepts)
    return linear(m, b)





def polyReg(X, Y, degree = 2):
    assert len(X) == len(Y)
    assert degree < len(X)
    vecY = np.transpose(np.atleast_2d(np.array(apply(Y, float))))
    matX = []
    for x in X:
        row = [float(x**i) for i in range(degree+1)]
        matX.append(row)
    matX = np.array(matX)
    inverse = np.linalg.inv(np.dot(matX.T, matX))
    parameters = np.dot(np.dot(inverse, matX.T), vecY)
    parameters = parameters.T.tolist()[0]
    ps = []
    for p in parameters[::-1]:
        ps.append(round(Decimal(p), 10))
    return polynomial(ps)

def mirrorReg(X, Y):
    assert len(X) == len(Y)
    X2 = [x ** 2 for x in X]
    lin = linReg(X2, Y)
    a = lin.a
    c = lin.b
    return polynomial([a, Decimal(0), c])


def sqrtReg(X, Y):
    assert len(X) == len(Y)
    half_lnX = [x.ln() / 2 for x in X]
    lnY = [y.ln() for y in Y]
    con = conReg(half_lnX, lnY)
    k = round(con.b.exp(), 10)
    return power(Decimal('0.5'), k)

def powReg(X, Y):
    assert len(X) == len(Y)
    lnX = [x.ln() for x in X]
    lnY = [y.ln() for y in Y]
    lin = linReg(lnX, lnY)
    a = round(lin.a, 10)
    b = round(lin.b.exp(), 10)
    return power(a, b)

def expReg(X, Y):
    lnY = [y.ln() for y in Y]
    lin = linReg(X, lnY)
    a = round(lin.a.exp(), 10)
    b = round(lin.b.exp(), 10)
    return exponential(a, b)

def logReg(X,Y):
    lnX = [x.ln() for x in X]
    lin = linReg(lnX, Y)
    a = round(lin.a, 10)
    b = round(lin.b, 10)
    return logarithmic(a, b)

def multReg(Y, *XS):
    assert len(XS[0]) == len(Y)
    vecY = np.transpose(np.atleast_2d(np.array(apply(Y, float))))
    matX = []
    for i in range(len(XS[0])):
        row = [1]
        row.extend([float(XS[j][i]) for j in range(len(XS))])
        matX.append(row)
    matX = np.array(matX)
    inverse = np.linalg.inv(np.dot(matX.T, matX))
    parameters = np.dot(np.dot(inverse, matX.T), vecY)
    parameters = parameters.T.tolist()[0]
    ps = []
    for p in parameters:
        ps.append(round(Decimal(p), 10))
    return multiLinear(ps)

def plotStuff(**kwargs):
    plt.figure()
    for name, data in kwargs.items():
        if isinstance(data, (list, tuple)):
            print(data)
            plt.plot(*data)

# Analyzers

def analyzeDataset(X):
    print('Values:')
    print(', '.join([str(x) for x in X]))
    print('\nMean:')
    print(mean(X))
    print('\nMedian:')
    print(median(X))
    print('\nMode:')
    print(mode(X)) 
    print('\nSum of values:')
    print(sum(X)) 
    print('\nMinimum value:')
    print(min(X)) 
    print('\nMaximum value:')
    print(max(X)) 
    print('\nVariation:')
    print(variation(X)) 
    print('\nMean absolute deviation')
    print(meanAbsoluteDeviation(X))
    print('\nVariance:')
    print(variance(X))
    print('\nStandard Deviation:')
    print(standardDeviation(X))

# Personal setup for excel file multiple regression

data = openpyxl.load_workbook(filename=r"C:\Users\Jacob\Desktop\Data.xlsx",data_only=True)
sheet = data.active

X1 = []
X2 = []
X3 = []
Y = []
for row in range(3,121):
    X1.append(sheet.cell(row=row, column=3).value)
    X2.append(sheet.cell(row=row, column=4).value)
    X3.append(sheet.cell(row=row, column=5).value)
    Y.append(sheet.cell(row=row, column=6).value)

X1 = apply(X1, Decimal)
X2 = apply(X2, Decimal)
X3 = apply(X3, Decimal)
Y = apply(Y, Decimal)


X2 = apply(X2, lambda x: x.sqrt())
X3 = apply(X3, lambda x: x**2)

model1 = multReg(Y, X1, X2, X3)
print(model1)
print(r2([X1,X2,X3], Y, model1))

Lg = []
Lgm = []
LgV = []
LgmV = []
T = []
for row in range(3,121):
    Lg.append(sheet.cell(row=row, column=13).value)
    Lgm.append(sheet.cell(row=row, column=14).value)
    LgV.append(sheet.cell(row=row, column=15).value)
    LgmV.append(sheet.cell(row=row, column=16).value)
    T.append(sheet.cell(row=row, column=17).value)

Lg = apply(Lg, Decimal)
Lgm = apply(Lgm, Decimal)
LgV = apply(LgV, Decimal)
LgmV = apply(LgmV, Decimal)
T = apply(T, Decimal)

model2 = multReg(T, Lg, Lgm, LgV, LgmV)
print(model2)
print(r2([Lg,Lgm,LgV, LgmV], T, model2))

if __name__ == '__main__':
    while True:
        exec(input())